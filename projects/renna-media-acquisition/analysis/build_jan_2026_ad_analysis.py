#!/usr/bin/env python3
"""Build a first-pass Jan 2026 Renna Media ad census.

This is intentionally conservative:
- verified Westfield rows are used as the source for known advertiser names/sizes
- all other papers are scanned for ad-like boxes containing phone numbers
- full-page creatives are grouped by visual hash so repeated image-only ads are not lost
- PSA/editorial/Renna-self candidates are retained but excluded from paid rollups
"""

from __future__ import annotations

import csv
import hashlib
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
PDF_ROOT = ROOT / "inputs" / "editions"
SOURCE_XLSX = ROOT / "analysis" / "jan-2026-ad-inventory.xlsx"
OUT_XLSX = ROOT / "analysis" / "jan-2026-ad-analysis.xlsx"
OUT_CSV = ROOT / "analysis" / "jan-2026-ad-analysis-inventory.csv"
MONTH = "2026-01"

PHONE_RE = re.compile(r"(?:\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4})")
URL_RE = re.compile(r"\b(?:www\.)?[a-z0-9][a-z0-9-]+\.(?:com|org|net|edu|us|info)\b", re.I)

SELF_PHONES = {"908-447-1295"}
RATE_SHEET_MARKERS = ("RENNA MEDIA NEWSPAPERS RATE SHEET", "Advertise in any number of towns")

SIZE_CANDIDATES = [
    ("1U", "3.25x2", 6.5, [(3.25, 2), (2, 3.25)]),
    ("2U", "6.5x2 or 3.25x4", 13, [(6.5, 2), (3.25, 4), (4, 3.25)]),
    ("4U", "6.5x4 or 3.25x8", 26, [(6.5, 4), (3.25, 8), (8, 3.25)]),
    ("6U", "6.5x6.25 or 10x4", 40, [(6.5, 6.25), (10, 4), (4, 10)]),
    ("9HP", "10x7 or 6.5x9", 65, [(10, 7), (6.5, 9), (9, 6.5)]),
    ("12U", "6.5x14", 91, [(6.5, 14), (14, 6.5)]),
    ("18FP", "10x14", 140, [(10, 14)]),
]


@dataclass
class Box:
    x0: float
    y0: float
    x1: float
    y1: float

    @property
    def width(self) -> float:
        return self.x1 - self.x0

    @property
    def height(self) -> float:
        return self.y1 - self.y0

    @property
    def area(self) -> float:
        return self.width * self.height

    def rect(self) -> fitz.Rect:
        return fitz.Rect(self.x0, self.y0, self.x1, self.y1)


def norm_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return value


def clean_text(value: str, limit: int = 220) -> str:
    value = re.sub(r"\s+", " ", value or "").strip()
    return value[:limit]


def title_from_slug(slug: str) -> str:
    specials = {"the-chathams": "The Chathams"}
    return specials.get(slug, slug.replace("-", " ").title())


def load_verified_rows():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Ad Inventory"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    phone_map = {}
    page_map = defaultdict(list)
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, r))
        if not row.get("paper_slug"):
            continue
        rows.append(row)
        page_map[int(row["page"])].append(row)
        phone = row.get("phone")
        if phone:
            phone_map[norm_phone(str(phone))] = row
    return rows, phone_map, page_map


def load_phone_categories():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    out = {}
    if "Phone Bundle (Auto)" not in wb.sheetnames:
        return out
    ws = wb["Phone Bundle (Auto)"]
    headers = None
    for r in ws.iter_rows(values_only=True):
        if r and r[0] == "Phone":
            headers = list(r)
            continue
        if not headers or not r or not r[0]:
            continue
        row = dict(zip(headers, r))
        out[norm_phone(str(row["Phone"]))] = row
    return out


def load_circulation():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Circulation"]
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r and r[0]:
            out[str(r[0])] = {"display": r[1], "circulation": r[2]}
    return out


def candidate_boxes(page: fitz.Page):
    boxes: list[Box] = []
    for drawing in page.get_drawings():
        rect = drawing.get("rect")
        if not rect:
            continue
        box = Box(rect.x0, rect.y0, rect.x1, rect.y1)
        if box.width < 60 or box.height < 40 or box.area < 5000:
            continue
        boxes.append(box)

    unique: list[Box] = []
    for box in sorted(boxes, key=lambda b: b.area, reverse=True):
        if any(
            abs(box.x0 - seen.x0) < 3
            and abs(box.y0 - seen.y0) < 3
            and abs(box.x1 - seen.x1) < 3
            and abs(box.y1 - seen.y1) < 3
            for seen in unique
        ):
            continue
        unique.append(box)
    return unique


def classify_size(box: Box, page_number: int, page_count: int):
    if box.width > 650 and box.height > 850:
        return ("BACK" if page_number == page_count else "18FP", "10x14", 140, "high")

    width_in = box.width / 72.0
    height_in = box.height / 72.0
    best = None
    for size_class, dims, sq_in, candidates in SIZE_CANDIDATES:
        for w, h in candidates:
            score = abs(width_in - w) / max(w, 0.1) + abs(height_in - h) / max(h, 0.1)
            if best is None or score < best[0]:
                best = (score, size_class, dims, sq_in)
    assert best is not None
    score, size_class, dims, sq_in = best
    confidence = "medium" if score < 0.35 else "low"
    return size_class, dims, sq_in, confidence


def position_for(box: Box, page_number: int, page_count: int, size_class: str):
    if size_class == "BACK":
        return "back_cover"
    if size_class == "18FP":
        return "full_page"
    if box.y0 < 120 and page_number == 1:
        return "front_box"
    if box.y0 < 250:
        return "top"
    if box.y0 > 720:
        return "bottom"
    return "inline"


def page_text(page: fitz.Page) -> str:
    return page.get_text("text") or ""


def is_rate_sheet(text: str) -> bool:
    upper = text.upper()
    return any(marker.upper() in upper for marker in RATE_SHEET_MARKERS)


def phash_clip(page: fitz.Page, box: Box) -> str:
    pix = page.get_pixmap(matrix=fitz.Matrix(0.12, 0.12), clip=box.rect(), alpha=False)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples).convert("L").resize((16, 16))
    pixels = list(img.getdata())
    avg = sum(pixels) / len(pixels)
    return "".join("1" if p > avg else "0" for p in pixels)


def hash_id(bits: str) -> str:
    return hashlib.sha1(bits.encode()).hexdigest()[:12]


def hamming(a: str, b: str) -> int:
    return sum(x != y for x, y in zip(a, b))


def build_known_creatives(page_map):
    known = {}
    pdf = PDF_ROOT / "westfield-monthly" / f"{MONTH}.pdf"
    doc = fitz.open(pdf)
    for page_no, rows in page_map.items():
        page = doc[page_no - 1]
        boxes = candidate_boxes(page)
        if not boxes:
            continue
        largest = max(boxes, key=lambda b: b.area)
        if largest.width < 650 or largest.height < 850:
            continue
        paid = [r for r in rows if not r.get("is_renna_self") and r.get("size_class") in {"18FP", "BACK"}]
        if not paid:
            continue
        row = paid[0]
        known[phash_clip(page, largest)] = row
    return known


def match_known_creative(hash_value: str, known):
    if hash_value in known:
        return known[hash_value], "high"
    nearest = None
    for known_hash, row in known.items():
        distance = hamming(hash_value, known_hash)
        if nearest is None or distance < nearest[0]:
            nearest = (distance, row)
    if nearest and nearest[0] <= 22:
        return nearest[1], "medium"
    return None, None


def best_name_from_context(context: str, phone: str):
    compact = clean_text(context, 160)
    if phone and phone in compact:
        before = compact.split(phone)[0].strip(" -•|")
        words = before.split()
        if words:
            return " ".join(words[-8:])
    return compact[:80] or f"Unknown advertiser {phone}"


def status_for(phone, phone_info, context, is_self):
    if is_self or phone in SELF_PHONES:
        return "renna_self"
    info = phone_info.get(phone)
    if info:
        category = str(info.get("Category") or "").lower()
        if category in {"paid", "psa", "renna_self", "editorial"}:
            return category
    lower = (context or "").lower()
    if any(k in lower for k in ["library", "parish", "chorus", "audition", "volunteer", "blood drive"]):
        return "psa"
    return "paid_candidate"


def choose_primary_phone(phones, phone_map, phone_info):
    for phone in phones:
        row = phone_map.get(phone)
        if row and not row.get("is_renna_self"):
            return phone
    for phone in phones:
        info = phone_info.get(phone)
        if info and str(info.get("Category") or "").lower() == "paid":
            return phone
    for phone in phones:
        if phone not in SELF_PHONES:
            return phone
    return phones[0] if phones else ""


def review_flag(status: str, confidence: str, advertiser: str, category: str) -> str:
    if status in {"renna_self", "psa", "editorial"}:
        return "no"
    if status == "paid_candidate":
        return "yes"
    if confidence == "low":
        return "yes"
    if category == "unknown" or str(advertiser).startswith("Unknown "):
        return "yes"
    return "no"


def write_sheet(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 55)
        ws.column_dimensions[col[0].column_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    verified_rows, phone_map, page_map = load_verified_rows()
    phone_info = load_phone_categories()
    circulation = load_circulation()
    known_creatives = build_known_creatives(page_map)

    inventory = []
    seen = set()

    # Preserve manually verified Westfield rows.
    for row in verified_rows:
        key = ("verified", row["paper_slug"], row["page"], row["advertiser_normalized"], row.get("phone"))
        seen.add(key)
        inventory.append(
            {
                "source": "verified_westfield",
                "status": "renna_self" if row.get("is_renna_self") else "paid",
                "paper_slug": row["paper_slug"],
                "paper_display": row["paper_display"],
                "circulation": row["circulation"],
                "month": row["month"],
                "page": row["page"],
                "advertiser_raw": row["advertiser_raw"],
                "advertiser_normalized": row["advertiser_normalized"],
                "category": row["category"],
                "size_class": row["size_class"],
                "size_dims": row["size_dims"],
                "size_sq_in": row["size_sq_in"],
                "position": row["position"],
                "phone": norm_phone(str(row["phone"])) if row.get("phone") else "",
                "website": row.get("website") or "",
                "confidence": row.get("confidence") or "high",
                "review_needed": "no",
                "notes": row.get("notes") or "",
            }
        )

    for pdf in sorted(PDF_ROOT.glob(f"*/{MONTH}.pdf")):
        slug = pdf.parent.name
        if slug == "westfield-monthly":
            continue
        doc = fitz.open(pdf)
        paper = circulation.get(slug, {"display": title_from_slug(slug), "circulation": ""})
        for idx, page in enumerate(doc, start=1):
            text = page_text(page)
            if is_rate_sheet(text):
                continue
            boxes = candidate_boxes(page)
            for box in boxes:
                clip_text = page.get_textbox(box.rect())
                phones = sorted({norm_phone(m.group(0)) for m in PHONE_RE.finditer(clip_text)})
                urls = sorted({m.group(0).lower() for m in URL_RE.finditer(clip_text)})
                size_class, size_dims, size_sq_in, size_conf = classify_size(box, idx, doc.page_count)
                pos = position_for(box, idx, doc.page_count, size_class)

                if phones:
                    phone = choose_primary_phone(phones, phone_map, phone_info)
                    info = phone_info.get(phone, {})
                    verified = phone_map.get(phone)
                    status = status_for(phone, phone_info, clip_text, phone in SELF_PHONES)
                    if status in {"editorial"}:
                        continue
                    if verified and not verified.get("is_renna_self"):
                        status = "paid"
                        adv_raw = verified["advertiser_raw"]
                        adv_norm = verified["advertiser_normalized"]
                        category = verified["category"]
                        size_class = verified["size_class"]
                        size_dims = verified["size_dims"]
                        size_sq_in = verified["size_sq_in"]
                        confidence = "medium"
                        notes = "Matched to verified Westfield advertiser by phone."
                    elif info:
                        adv_raw = info.get("Best-Guess Advertiser") or best_name_from_context(clip_text, phone)
                        adv_norm = adv_raw
                        category = info.get("Category") or status
                        confidence = "medium" if status == "paid" else "low"
                        notes = info.get("Notes") or "Matched by phone fingerprint from auto bundle sheet."
                    else:
                        adv_raw = best_name_from_context(clip_text, phone)
                        adv_norm = adv_raw
                        category = "unknown"
                        confidence = size_conf
                        notes = "Auto-detected phone inside ad-like rectangle."
                    key = (slug, idx, ";".join(phones), round(box.x0), round(box.y0), round(box.x1), round(box.y1))
                    if key in seen:
                        continue
                    seen.add(key)
                    inventory.append(
                        {
                            "source": "auto_phone_box",
                            "status": "paid" if status == "paid" else status,
                            "paper_slug": slug,
                            "paper_display": paper["display"],
                            "circulation": paper["circulation"],
                            "month": MONTH,
                            "page": idx,
                            "advertiser_raw": adv_raw,
                            "advertiser_normalized": adv_norm,
                            "category": category,
                            "size_class": size_class,
                            "size_dims": size_dims,
                            "size_sq_in": size_sq_in,
                            "position": pos,
                            "phone": "; ".join(phones),
                            "website": "; ".join(urls),
                            "confidence": confidence,
                            "review_needed": review_flag(status, confidence, adv_norm, category),
                            "notes": notes,
                        }
                    )
                elif size_class in {"18FP", "BACK"}:
                    hash_value = phash_clip(page, box)
                    known, conf = match_known_creative(hash_value, known_creatives)
                    if known:
                        adv_raw = known["advertiser_raw"]
                        adv_norm = known["advertiser_normalized"]
                        category = known["category"]
                        confidence = conf
                        notes = f"Matched full-page creative hash from verified Westfield page {known['page']}."
                        status = "paid"
                        phone = norm_phone(str(known["phone"])) if known.get("phone") else ""
                        website = known.get("website") or ""
                    else:
                        creative_id = hash_id(hash_value)
                        adv_raw = f"Unknown full-page creative {creative_id}"
                        adv_norm = adv_raw
                        category = "unknown"
                        confidence = "low"
                        notes = "Full-page/back-cover ad-like creative with no extractable phone/text. Needs visual review."
                        status = "paid_candidate"
                        phone = ""
                        website = ""
                    key = (slug, idx, hash_id(hash_value), size_class)
                    if key in seen:
                        continue
                    seen.add(key)
                    inventory.append(
                        {
                            "source": "auto_full_page_hash",
                            "status": status,
                            "paper_slug": slug,
                            "paper_display": paper["display"],
                            "circulation": paper["circulation"],
                            "month": MONTH,
                            "page": idx,
                            "advertiser_raw": adv_raw,
                            "advertiser_normalized": adv_norm,
                            "category": category,
                            "size_class": size_class,
                            "size_dims": "10x14",
                            "size_sq_in": 140,
                            "position": pos,
                            "phone": phone,
                            "website": website,
                            "confidence": confidence,
                            "review_needed": review_flag(status, confidence, adv_norm, category),
                            "notes": notes,
                        }
                    )

    inventory.sort(key=lambda r: (r["paper_slug"], int(r["page"]), r["advertiser_normalized"], r["phone"]))

    paid_statuses = {"paid", "paid_candidate"}
    paid_rows = [r for r in inventory if r["status"] in paid_statuses and r["category"] != "renna_self"]
    advertiser_papers = defaultdict(set)
    advertiser_sizes = defaultdict(Counter)
    advertiser_rows = defaultdict(int)
    advertiser_conf = defaultdict(Counter)
    for r in paid_rows:
        adv = r["advertiser_normalized"]
        advertiser_papers[adv].add(r["paper_slug"])
        advertiser_sizes[adv][r["size_class"]] += 1
        advertiser_rows[adv] += 1
        advertiser_conf[adv][r["confidence"]] += 1

    rollup = []
    for adv, papers in sorted(advertiser_papers.items(), key=lambda item: (-len(item[1]), item[0])):
        sizes = advertiser_sizes[adv]
        rollup.append(
            {
                "advertiser_normalized": adv,
                "num_papers": len(papers),
                "num_ad_rows": advertiser_rows[adv],
                "primary_size": sizes.most_common(1)[0][0],
                "all_sizes": "; ".join(f"{k}:{v}" for k, v in sizes.most_common()),
                "papers": "; ".join(sorted(papers)),
                "confidence_mix": "; ".join(f"{k}:{v}" for k, v in advertiser_conf[adv].most_common()),
            }
        )

    paper_summary = []
    for slug in sorted({r["paper_slug"] for r in inventory}):
        rows = [r for r in inventory if r["paper_slug"] == slug]
        paid = [r for r in rows if r["status"] in paid_statuses and r["category"] != "renna_self"]
        paper_summary.append(
            {
                "paper_slug": slug,
                "paper_display": rows[0]["paper_display"],
                "paid_ad_rows": len(paid),
                "unique_paid_advertisers": len({r["advertiser_normalized"] for r in paid}),
                "high_confidence_rows": sum(1 for r in paid if r["confidence"] == "high"),
                "review_needed_rows": sum(1 for r in paid if r["review_needed"] == "yes"),
                "full_page_or_back_rows": sum(1 for r in paid if r["size_class"] in {"18FP", "BACK"}),
            }
        )

    papers = sorted({r["paper_slug"] for r in inventory})
    matrix_rows = []
    by_adv_paper = defaultdict(list)
    for r in paid_rows:
        by_adv_paper[(r["advertiser_normalized"], r["paper_slug"])].append(r["size_class"])
    for row in rollup:
        matrix_row = {"advertiser_normalized": row["advertiser_normalized"], "num_papers": row["num_papers"]}
        for p in papers:
            vals = by_adv_paper.get((row["advertiser_normalized"], p), [])
            matrix_row[p] = "; ".join(vals)
        matrix_rows.append(matrix_row)

    needs_review = [r for r in inventory if r["review_needed"] == "yes" or r["status"] == "paid_candidate"]

    with OUT_CSV.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(inventory[0].keys()))
        writer.writeheader()
        writer.writerows(inventory)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    readme = wb.create_sheet("README")
    readme_rows = [
        ("Renna Media Jan 2026 Ad Analysis", ""),
        ("Purpose", "Track advertisers, ad size, and paper distribution across all 30 Jan 2026 papers."),
        ("Important", "This is a first-pass census. Westfield is manually verified; other papers are auto-scanned and flagged for review where confidence is not high."),
        ("Paid Count Rule", "Rollups include status=paid and status=paid_candidate, excluding Renna self-promo. PSA/editorial rows are retained in the inventory but excluded."),
        ("Generated From", str(SOURCE_XLSX.relative_to(ROOT))),
        ("Output CSV", str(OUT_CSV.relative_to(ROOT))),
    ]
    for r in readme_rows:
        readme.append(r)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 110
    for cell in readme["A"]:
        cell.font = Font(bold=True)

    write_sheet(wb.create_sheet("Ad Inventory"), inventory)
    write_sheet(wb.create_sheet("Advertiser Rollup"), rollup)
    write_sheet(wb.create_sheet("Advertiser x Paper"), matrix_rows)
    write_sheet(wb.create_sheet("Paper Summary"), paper_summary)
    write_sheet(wb.create_sheet("Needs Review"), needs_review)

    bundle_counts = Counter()
    for row in rollup:
        n = row["num_papers"]
        if n == 1:
            bucket = "1 paper"
        elif n <= 3:
            bucket = "2-3 papers"
        elif n <= 6:
            bucket = "4-6 papers"
        elif n <= 12:
            bucket = "7-12 papers"
        elif n <= 20:
            bucket = "13-20 papers"
        else:
            bucket = "21-30 papers"
        bundle_counts[bucket] += 1
    bundle_rows = [
        {"bundle_bucket": bucket, "advertisers": bundle_counts[bucket]}
        for bucket in ["1 paper", "2-3 papers", "4-6 papers", "7-12 papers", "13-20 papers", "21-30 papers"]
    ]
    write_sheet(wb.create_sheet("Bundle Summary"), bundle_rows)

    wb.save(OUT_XLSX)

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Inventory rows: {len(inventory)}")
    print(f"Paid/potential paid rows: {len(paid_rows)}")
    print(f"Unique paid/potential advertisers: {len(rollup)}")
    print(f"Needs review rows: {len(needs_review)}")


if __name__ == "__main__":
    main()
