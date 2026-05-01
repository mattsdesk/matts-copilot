#!/usr/bin/env python3
"""Fast 2024-2025 retention analysis using phone-number fingerprints.

This is the cleanest automated signal for retention because phone numbers stay
stable even when ad creative, copy, and size change. It does not try to infer ad
size from every historical PDF; use it as the retention/cross-paper baseline.
"""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


ROOT = Path(__file__).resolve().parents[1]
PDF_ROOT = ROOT / "inputs" / "editions"
SOURCE_XLSX = ROOT / "analysis" / "jan-2026-ad-inventory.xlsx"
OUT_XLSX = ROOT / "analysis" / "phone-retention-2024-2025.xlsx"
OUT_CSV = ROOT / "analysis" / "phone-retention-2024-2025-observations.csv"
SUMMARY_MD = ROOT / "analysis" / "phone-retention-2024-2025-summary.md"

MONTHS = [f"{year}-{month:02d}" for year in (2024, 2025) for month in range(1, 13)]
PHONE_RE = re.compile(r"(?:\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4})")
URL_RE = re.compile(r"\b(?:www\.)?[a-z0-9][a-z0-9-]+\.(?:com|org|net|edu|us|info)\b", re.I)
SELF_PHONES = {"908-447-1295", "908-418-5586"}
RATE_SHEET_MARKERS = ("RENNA MEDIA NEWSPAPERS RATE SHEET", "Advertise in any number of towns")
PSA_MARKERS = (
    "library",
    "parish",
    "chorus",
    "audition",
    "volunteer",
    "blood drive",
    "senior center",
    "scout",
    "elks",
    "rotary",
    "free event",
)


def norm_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return value


def clean_text(value: str, limit: int = 220) -> str:
    return re.sub(r"\s+", " ", value or "").strip()[:limit]


def title_from_slug(slug: str) -> str:
    return {"the-chathams": "The Chathams"}.get(slug, slug.replace("-", " ").title())


def load_circulation():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Circulation"]
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r and r[0]:
            out[str(r[0])] = {"display": r[1], "circulation": r[2]}
    return out


def load_known_names():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    known = {}
    ws = wb["Ad Inventory"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, r))
        phone = row.get("phone")
        if phone and not row.get("is_renna_self"):
            known[norm_phone(str(phone))] = {
                "name": row.get("advertiser_normalized") or row.get("advertiser_raw"),
                "category": row.get("category") or "paid",
            }
    if "Phone Bundle (Auto)" in wb.sheetnames:
        ws = wb["Phone Bundle (Auto)"]
        headers = None
        for r in ws.iter_rows(values_only=True):
            if r and r[0] == "Phone":
                headers = list(r)
                continue
            if not headers or not r or not r[0]:
                continue
            row = dict(zip(headers, r))
            phone = norm_phone(str(row["Phone"]))
            category = str(row.get("Category") or "").lower()
            if category == "paid":
                known.setdefault(
                    phone,
                    {
                        "name": row.get("Best-Guess Advertiser") or phone,
                        "category": "paid",
                    },
                )
    return known


def status_for(phone: str, context: str, known) -> str:
    if phone in SELF_PHONES:
        return "renna_self"
    if phone in known:
        return "paid"
    lower = context.lower()
    if any(marker in lower for marker in PSA_MARKERS):
        return "psa"
    return "paid_candidate"


def name_from_context(phone: str, context: str, known) -> str:
    if phone in known:
        return known[phone]["name"]
    compact = clean_text(context, 180)
    if phone in compact:
        before = compact.split(phone)[0].strip(" -•|")
        words = before.split()
        if words:
            return " ".join(words[-8:])
    urls = sorted({m.group(0).lower() for m in URL_RE.finditer(context or "")})
    return urls[0] if urls else f"Unknown phone {phone}"


def context_window(text: str, start: int, end: int, span: int = 140) -> str:
    return clean_text(text[max(0, start - span) : min(len(text), end + span)], 300)


def scan():
    known = load_known_names()
    circulation = load_circulation()
    rows = []
    coverage = []

    for month in MONTHS:
        for pdf in sorted(PDF_ROOT.glob(f"*/{month}.pdf")):
            slug = pdf.parent.name
            paper = circulation.get(slug, {"display": title_from_slug(slug), "circulation": ""})
            doc = fitz.open(pdf)
            issue_count = 0
            for page_no, page in enumerate(doc, start=1):
                text = page.get_text("text") or ""
                upper = text.upper()
                if any(marker.upper() in upper for marker in RATE_SHEET_MARKERS):
                    continue
                seen_on_page = set()
                for match in PHONE_RE.finditer(text):
                    phone = norm_phone(match.group(0))
                    if phone in seen_on_page:
                        continue
                    seen_on_page.add(phone)
                    context = context_window(text, match.start(), match.end())
                    status = status_for(phone, context, known)
                    if status == "renna_self":
                        continue
                    rows.append(
                        {
                            "month": month,
                            "year": int(month[:4]),
                            "paper_slug": slug,
                            "paper_display": paper["display"],
                            "circulation": paper["circulation"],
                            "page": page_no,
                            "phone": phone,
                            "status": status,
                            "advertiser_name": name_from_context(phone, context, known),
                            "website": "; ".join(sorted({m.group(0).lower() for m in URL_RE.finditer(context)})),
                            "context": context,
                            "source_pdf": str(pdf.relative_to(ROOT)),
                        }
                    )
                    issue_count += 1
            coverage.append(
                {
                    "month": month,
                    "paper_slug": slug,
                    "paper_display": paper["display"],
                    "pages": doc.page_count,
                    "phone_observations": issue_count,
                    "source_pdf": str(pdf.relative_to(ROOT)),
                }
            )
    return rows, coverage


def bucket(n: int) -> str:
    if n == 1:
        return "1 paper"
    if n <= 3:
        return "2-3 papers"
    if n <= 6:
        return "4-6 papers"
    if n <= 12:
        return "7-12 papers"
    if n <= 20:
        return "13-20 papers"
    return "21-30 papers"


def build_rollups(rows, coverage):
    paid = [r for r in rows if r["status"] in {"paid", "paid_candidate"}]
    by_phone = defaultdict(list)
    by_phone_month = defaultdict(list)
    by_month = defaultdict(list)
    for row in paid:
        by_phone[row["phone"]].append(row)
        by_phone_month[(row["phone"], row["month"])].append(row)
        by_month[row["month"]].append(row)

    advertiser_summary = []
    for phone, group in sorted(by_phone.items(), key=lambda item: (-len({r["month"] for r in item[1]}), item[0])):
        months = sorted({r["month"] for r in group})
        papers = sorted({r["paper_slug"] for r in group})
        names = Counter(r["advertiser_name"] for r in group)
        month_paper_counts = [len({r["paper_slug"] for r in by_phone_month[(phone, m)]}) for m in months]
        active_2024 = sum(1 for m in months if m.startswith("2024-"))
        active_2025 = sum(1 for m in months if m.startswith("2025-"))
        advertiser_summary.append(
            {
                "phone": phone,
                "advertiser_name": names.most_common(1)[0][0],
                "active_months": len(months),
                "active_2024_months": active_2024,
                "active_2025_months": active_2025,
                "first_seen": months[0],
                "last_seen": months[-1],
                "retained_2024_to_2025": "yes" if active_2024 and active_2025 else "no",
                "max_papers_in_month": max(month_paper_counts),
                "avg_papers_when_active": round(sum(month_paper_counts) / len(month_paper_counts), 2),
                "paper_bucket_at_peak": bucket(max(month_paper_counts)),
                "total_observations": len(group),
                "all_papers": "; ".join(papers),
            }
        )

    advertiser_month = []
    for (phone, month), group in sorted(by_phone_month.items(), key=lambda item: (item[0][1], item[0][0])):
        papers = sorted({r["paper_slug"] for r in group})
        names = Counter(r["advertiser_name"] for r in group)
        advertiser_month.append(
            {
                "phone": phone,
                "advertiser_name": names.most_common(1)[0][0],
                "month": month,
                "num_papers": len(papers),
                "paper_bucket": bucket(len(papers)),
                "observations": len(group),
                "papers": "; ".join(papers),
            }
        )

    monthly_summary = []
    for i, month in enumerate(MONTHS):
        active = {r["phone"] for r in by_month.get(month, [])}
        next_retained = ""
        if i < len(MONTHS) - 1 and active:
            next_active = {r["phone"] for r in by_month.get(MONTHS[i + 1], [])}
            next_retained = len(active & next_active)
        paper_counts = [len({r["paper_slug"] for r in by_phone_month[(phone, month)]}) for phone in active]
        monthly_summary.append(
            {
                "month": month,
                "issues_available": len([c for c in coverage if c["month"] == month]),
                "active_paid_phones": len(active),
                "observations": len(by_month.get(month, [])),
                "retained_next_month": next_retained,
                "next_month_retention_rate": round(next_retained / len(active), 3) if isinstance(next_retained, int) and active else "",
                "avg_papers_per_active_phone": round(sum(paper_counts) / len(paper_counts), 2) if paper_counts else 0,
                "phones_1_paper": sum(1 for n in paper_counts if n == 1),
                "phones_2_3_papers": sum(1 for n in paper_counts if 2 <= n <= 3),
                "phones_4_6_papers": sum(1 for n in paper_counts if 4 <= n <= 6),
                "phones_7_12_papers": sum(1 for n in paper_counts if 7 <= n <= 12),
                "phones_13_plus_papers": sum(1 for n in paper_counts if n >= 13),
            }
        )

    retention_matrix = []
    for row in advertiser_summary:
        phone = row["phone"]
        out = {
            "phone": phone,
            "advertiser_name": row["advertiser_name"],
            "active_months": row["active_months"],
            "retained_2024_to_2025": row["retained_2024_to_2025"],
            "max_papers_in_month": row["max_papers_in_month"],
        }
        for month in MONTHS:
            out[month] = len({r["paper_slug"] for r in by_phone_month.get((phone, month), [])})
        retention_matrix.append(out)

    cohorts = []
    by_first = defaultdict(list)
    for row in advertiser_summary:
        by_first[row["first_seen"]].append(row["phone"])
    for i, month in enumerate(MONTHS):
        cohort = by_first.get(month, [])
        out = {"cohort_month": month, "new_paid_phones": len(cohort)}
        for offset in range(min(12, len(MONTHS) - i)):
            target = MONTHS[i + offset]
            active_count = sum(1 for phone in cohort if (phone, target) in by_phone_month)
            out[f"m+{offset}"] = active_count
            out[f"m+{offset}_rate"] = round(active_count / len(cohort), 3) if cohort else ""
        cohorts.append(out)

    paper_month = []
    paid_by_paper_month = defaultdict(list)
    for row in paid:
        paid_by_paper_month[(row["paper_slug"], row["month"])].append(row)
    for row in coverage:
        group = paid_by_paper_month.get((row["paper_slug"], row["month"]), [])
        paper_month.append(
            {
                **row,
                "paid_phone_observations": len(group),
                "unique_paid_phones": len({r["phone"] for r in group}),
            }
        )

    return {
        "observations": rows,
        "paid": paid,
        "advertiser_summary": advertiser_summary,
        "advertiser_month": advertiser_month,
        "monthly_summary": monthly_summary,
        "retention_matrix": retention_matrix,
        "cohorts": cohorts,
        "paper_month": paper_month,
        "coverage": coverage,
    }


def write_sheet(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([clean_cell(row.get(h)) for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 70)
        ws.column_dimensions[col[0].column_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def clean_cell(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def write_outputs(data):
    with OUT_CSV.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(data["observations"][0].keys()))
        writer.writeheader()
        writer.writerows(data["observations"])

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    for row in [
        ("Renna Media Phone Retention 2024-2025", ""),
        ("Scope", "All available 2024 and 2025 PDFs under inputs/editions."),
        ("Method", "Phone-number fingerprints extracted from page text, excluding rate-sheet pages and Renna self phone."),
        ("Best Use", "Retention and cross-paper buying behavior. Not a final ad-size or revenue analysis."),
        ("Caveat", "Some PSAs/community notices may remain as paid candidates until reconciled with invoice data."),
    ]:
        readme.append(row)
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 110
    for cell in readme["A"]:
        cell.font = Font(bold=True)

    for name, rows in [
        ("Monthly Summary", data["monthly_summary"]),
        ("Advertiser Summary", data["advertiser_summary"]),
        ("Advertiser Month", data["advertiser_month"]),
        ("Retention Matrix", data["retention_matrix"]),
        ("Cohorts", data["cohorts"]),
        ("Paper Month", data["paper_month"]),
        ("Observations", data["observations"]),
    ]:
        write_sheet(wb.create_sheet(name), rows)
    wb.save(OUT_XLSX)


def write_summary(data):
    adv = data["advertiser_summary"]
    monthly = data["monthly_summary"]
    retained = [r for r in adv if r["retained_2024_to_2025"] == "yes"]
    active_12 = [r for r in adv if r["active_months"] >= 12]
    active_18 = [r for r in adv if r["active_months"] >= 18]
    retention_rates = [r["next_month_retention_rate"] for r in monthly if isinstance(r["next_month_retention_rate"], float)]
    top_reach = sorted(adv, key=lambda r: (-r["max_papers_in_month"], -r["active_months"]))[:15]
    top_duration = sorted(adv, key=lambda r: (-r["active_months"], -r["max_papers_in_month"]))[:15]

    lines = [
        "# Phone Retention Analysis - 2024 and 2025",
        "",
        "## Outputs",
        "",
        "- `analysis/phone-retention-2024-2025.xlsx`",
        "- `analysis/phone-retention-2024-2025-observations.csv`",
        "- `analysis/build_2024_2025_phone_retention.py`",
        "",
        "## Headline Metrics",
        "",
        f"- Issues scanned: {len(data['coverage'])}",
        f"- Paid/potential-paid phone observations: {len(data['paid'])}",
        f"- Paid/potential-paid phone fingerprints: {len(adv)}",
        f"- Seen in both 2024 and 2025: {len(retained)}",
        f"- Active 12+ months: {len(active_12)}",
        f"- Active 18+ months: {len(active_18)}",
        f"- Average next-month retention: {round(sum(retention_rates) / len(retention_rates), 3) if retention_rates else 'n/a'}",
        "",
        "## Top Cross-Paper Phone Fingerprints",
        "",
        "| Advertiser / Phone | Active Months | Max Papers In Month | Avg Papers When Active | Peak Bucket |",
        "|---|---:|---:|---:|---|",
    ]
    for row in top_reach:
        lines.append(
            f"| {row['advertiser_name']} / {row['phone']} | {row['active_months']} | {row['max_papers_in_month']} | {row['avg_papers_when_active']} | {row['paper_bucket_at_peak']} |"
        )
    lines.extend(["", "## Longest-Running Phone Fingerprints", "", "| Advertiser / Phone | Active Months | Max Papers In Month | First Seen | Last Seen |", "|---|---:|---:|---|---|"])
    for row in top_duration:
        lines.append(
            f"| {row['advertiser_name']} / {row['phone']} | {row['active_months']} | {row['max_papers_in_month']} | {row['first_seen']} | {row['last_seen']} |"
        )
    lines.extend(
        [
            "",
            "## Caveats",
            "",
            "- This is a phone-fingerprint view, not a final customer ledger.",
            "- It is strong for retention and cross-paper purchasing, but weak for ad size and revenue.",
            "- Phone numbers can over-split one customer if an ad uses multiple numbers, or under-split if one agency/office number represents multiple advertisers.",
            "- Reconcile against Joe's invoice/customer export before using this in valuation.",
        ]
    )
    SUMMARY_MD.write_text("\n".join(lines) + "\n")


def main():
    rows, coverage = scan()
    data = build_rollups(rows, coverage)
    write_outputs(data)
    write_summary(data)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"Issues scanned: {len(coverage)}")
    print(f"Observations: {len(rows)}")
    print(f"Paid/potential paid observations: {len(data['paid'])}")
    print(f"Paid/potential paid phone fingerprints: {len(data['advertiser_summary'])}")


if __name__ == "__main__":
    main()
