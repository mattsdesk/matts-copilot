#!/usr/bin/env python3
"""Build a 2024-2025 advertiser retention and cross-paper analysis.

The output is a first-pass operating view, not a hand-audited invoice ledger.
It uses ad-like PDF boxes, phone fingerprints, and repeated full-page creative
hashes to estimate advertiser activity by month and paper.
"""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from build_jan_2026_ad_analysis import (
    MONTH,
    PDF_ROOT,
    ROOT,
    PHONE_RE,
    URL_RE,
    candidate_boxes,
    classify_size,
    clean_text,
    hash_id,
    is_rate_sheet,
    load_circulation,
    load_phone_categories,
    load_verified_rows,
    norm_phone,
    page_text,
    phash_clip,
    position_for,
    review_flag,
    title_from_slug,
)

SELF_PHONES = {"908-447-1295", "908-418-5586"}


OUT_XLSX = ROOT / "analysis" / "advertiser-retention-2024-2025.xlsx"
OUT_CSV = ROOT / "analysis" / "advertiser-retention-2024-2025-placements.csv"
SUMMARY_MD = ROOT / "analysis" / "advertiser-retention-2024-2025-summary.md"

MONTHS = [f"{year}-{month:02d}" for year in (2024, 2025) for month in range(1, 13)]

PSA_MARKERS = [
    "library",
    "parish",
    "chorus",
    "audition",
    "volunteer",
    "blood drive",
    "church",
    "senior center",
    "scouts",
    "scout",
    "elks",
    "rotary",
    "fundraiser",
    "free event",
    "donation",
]


def write_sheet(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 65)
        ws.column_dimensions[col[0].column_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def best_name_from_context(context: str, phone: str) -> str:
    compact = clean_text(context, 180)
    if phone and phone in compact:
        before = compact.split(phone)[0].strip(" -•|")
        words = before.split()
        if words:
            return " ".join(words[-8:])
    urls = sorted({m.group(0).lower() for m in URL_RE.finditer(context or "")})
    if urls:
        return urls[0]
    return compact[:80] or f"Unknown advertiser {phone}"


def choose_primary_phone(phones, phone_map, phone_info):
    for phone in phones:
        row = phone_map.get(phone)
        if row and not row.get("is_renna_self"):
            return phone
    for phone in phones:
        info = phone_info.get(phone)
        if info and str(info.get("Category") or "").lower() == "paid":
            return phone
    for phone in phones:
        if phone not in SELF_PHONES:
            return phone
    return phones[0] if phones else ""


def status_for_context(phone: str, context: str, phone_info) -> str:
    if phone in SELF_PHONES:
        return "renna_self"
    info = phone_info.get(phone)
    if info:
        category = str(info.get("Category") or "").lower()
        if category in {"paid", "psa", "renna_self", "editorial"}:
            return category
    lower = (context or "").lower()
    if any(marker in lower for marker in PSA_MARKERS):
        return "psa"
    return "paid_candidate"


def load_known_advertisers():
    verified_rows, phone_map, _ = load_verified_rows()
    phone_info = load_phone_categories()

    known = {}
    for phone, row in phone_map.items():
        if not row.get("is_renna_self"):
            known[phone] = {
                "advertiser_raw": row["advertiser_raw"],
                "advertiser_normalized": row["advertiser_normalized"],
                "category": row["category"],
            }
    for phone, row in phone_info.items():
        category = str(row.get("Category") or "").lower()
        if category == "paid":
            name = row.get("Best-Guess Advertiser") or phone
            known.setdefault(
                phone,
                {
                    "advertiser_raw": name,
                    "advertiser_normalized": name,
                    "category": "paid",
                },
            )
    return known, phone_map, phone_info


def issue_paths():
    for month in MONTHS:
        for pdf in sorted(PDF_ROOT.glob(f"*/{month}.pdf")):
            yield month, pdf.parent.name, pdf


def scan():
    circulation = load_circulation()
    known, phone_map, phone_info = load_known_advertisers()
    rows = []
    coverage = []

    for month, slug, pdf in issue_paths():
        import fitz

        doc = fitz.open(pdf)
        paper = circulation.get(slug, {"display": title_from_slug(slug), "circulation": ""})
        issue_rows = 0

        for page_no, page in enumerate(doc, start=1):
            text = page_text(page)
            if is_rate_sheet(text):
                continue

            for box in candidate_boxes(page):
                clip_text = page.get_textbox(box.rect())
                phones = sorted({norm_phone(m.group(0)) for m in PHONE_RE.finditer(clip_text)})
                urls = sorted({m.group(0).lower() for m in URL_RE.finditer(clip_text)})
                size_class, size_dims, size_sq_in, size_conf = classify_size(box, page_no, doc.page_count)
                position = position_for(box, page_no, doc.page_count, size_class)

                if phones:
                    phone = choose_primary_phone(phones, phone_map, phone_info)
                    status = status_for_context(phone, clip_text, phone_info)
                    if status == "editorial":
                        continue
                    matched = known.get(phone)
                    if matched:
                        advertiser_raw = matched["advertiser_raw"]
                        advertiser_normalized = matched["advertiser_normalized"]
                        category = matched["category"]
                        confidence = "medium"
                    else:
                        advertiser_raw = best_name_from_context(clip_text, phone)
                        advertiser_normalized = advertiser_raw
                        category = status
                        confidence = size_conf
                    fingerprint = f"phone:{phone}"
                    source = "phone_box"
                    notes = "Phone fingerprint from ad-like box."
                elif size_class in {"18FP", "BACK"}:
                    bits = phash_clip(page, box)
                    creative_id = hash_id(bits)
                    advertiser_raw = f"Unknown full-page creative {creative_id}"
                    advertiser_normalized = advertiser_raw
                    category = "unknown"
                    confidence = "low"
                    status = "paid_candidate"
                    fingerprint = f"creative:{creative_id}"
                    source = "full_page_hash"
                    phone = ""
                    notes = "Image-only full-page/back-cover creative. Retention only tracks this repeated creative, not legal customer identity."
                else:
                    continue

                review = review_flag(status, confidence, advertiser_normalized, category)
                rows.append(
                    {
                        "month": month,
                        "year": int(month[:4]),
                        "paper_slug": slug,
                        "paper_display": paper["display"],
                        "circulation": paper["circulation"],
                        "page": page_no,
                        "source": source,
                        "status": "paid" if status == "paid" else status,
                        "advertiser_fingerprint": fingerprint,
                        "advertiser_raw": advertiser_raw,
                        "advertiser_normalized": advertiser_normalized,
                        "category": category,
                        "size_class": size_class,
                        "size_dims": size_dims,
                        "size_sq_in": size_sq_in,
                        "position": position,
                        "phone": "; ".join(phones) if phones else "",
                        "website": "; ".join(urls) if urls else "",
                        "confidence": confidence,
                        "review_needed": review,
                        "notes": notes,
                    }
                )
                issue_rows += 1

        coverage.append(
            {
                "month": month,
                "paper_slug": slug,
                "paper_display": paper["display"],
                "pages": doc.page_count,
                "inventory_rows": issue_rows,
                "source_pdf": str(pdf.relative_to(ROOT)),
            }
        )

    return rows, coverage


def bucket_papers(n: int) -> str:
    if n == 1:
        return "1 paper"
    if n <= 3:
        return "2-3 papers"
    if n <= 6:
        return "4-6 papers"
    if n <= 12:
        return "7-12 papers"
    if n <= 20:
        return "13-20 papers"
    return "21-30 papers"


def build_rollups(rows, coverage):
    paid_rows = [
        r
        for r in rows
        if r["status"] in {"paid", "paid_candidate"}
        and r["category"] not in {"renna_self", "psa", "editorial"}
    ]

    by_adv_month = defaultdict(list)
    by_adv = defaultdict(list)
    by_month = defaultdict(list)
    by_paper_month = defaultdict(list)
    for row in paid_rows:
        by_adv_month[(row["advertiser_fingerprint"], row["month"])].append(row)
        by_adv[row["advertiser_fingerprint"]].append(row)
        by_month[row["month"]].append(row)
        by_paper_month[(row["paper_slug"], row["month"])].append(row)

    advertiser_month = []
    for (fingerprint, month), group in sorted(by_adv_month.items(), key=lambda item: (item[0][1], item[0][0])):
        papers = sorted({r["paper_slug"] for r in group})
        sizes = Counter(r["size_class"] for r in group)
        names = Counter(r["advertiser_normalized"] for r in group)
        advertiser_month.append(
            {
                "advertiser_fingerprint": fingerprint,
                "advertiser_normalized": names.most_common(1)[0][0],
                "month": month,
                "active": 1,
                "num_papers": len(papers),
                "num_ad_rows": len(group),
                "paper_bucket": bucket_papers(len(papers)),
                "primary_size": sizes.most_common(1)[0][0],
                "all_sizes": "; ".join(f"{k}:{v}" for k, v in sizes.most_common()),
                "papers": "; ".join(papers),
                "review_rows": sum(1 for r in group if r["review_needed"] == "yes"),
            }
        )

    advertiser_summary = []
    for fingerprint, group in sorted(by_adv.items(), key=lambda item: (-len({r["month"] for r in item[1]}), item[0])):
        months = sorted({r["month"] for r in group})
        papers = sorted({r["paper_slug"] for r in group})
        names = Counter(r["advertiser_normalized"] for r in group)
        sizes = Counter(r["size_class"] for r in group)
        month_paper_counts = [len({r["paper_slug"] for r in by_adv_month[(fingerprint, month)]}) for month in months]
        active_2024 = sum(1 for m in months if m.startswith("2024-"))
        active_2025 = sum(1 for m in months if m.startswith("2025-"))
        advertiser_summary.append(
            {
                "advertiser_fingerprint": fingerprint,
                "advertiser_normalized": names.most_common(1)[0][0],
                "active_months": len(months),
                "active_2024_months": active_2024,
                "active_2025_months": active_2025,
                "first_seen": months[0],
                "last_seen": months[-1],
                "retained_2024_to_2025": "yes" if active_2024 and active_2025 else "no",
                "max_papers_in_month": max(month_paper_counts),
                "avg_papers_when_active": round(sum(month_paper_counts) / len(month_paper_counts), 2),
                "total_ad_rows": len(group),
                "primary_size": sizes.most_common(1)[0][0],
                "all_sizes": "; ".join(f"{k}:{v}" for k, v in sizes.most_common()),
                "all_papers": "; ".join(papers),
                "review_rows": sum(1 for r in group if r["review_needed"] == "yes"),
            }
        )

    monthly_summary = []
    for month in MONTHS:
        group = by_month.get(month, [])
        active = {r["advertiser_fingerprint"] for r in group}
        retained_next = None
        if month != MONTHS[-1]:
            next_month = MONTHS[MONTHS.index(month) + 1]
            next_active = {r["advertiser_fingerprint"] for r in by_month.get(next_month, [])}
            retained_next = len(active & next_active)
        paper_counts = [
            len({r["paper_slug"] for r in by_adv_month[(fp, month)]})
            for fp in active
        ]
        monthly_summary.append(
            {
                "month": month,
                "issues_available": len([c for c in coverage if c["month"] == month]),
                "paid_ad_rows": len(group),
                "active_advertisers": len(active),
                "retained_next_month": retained_next,
                "next_month_retention_rate": round(retained_next / len(active), 3) if active and retained_next is not None else "",
                "avg_papers_per_active_advertiser": round(sum(paper_counts) / len(paper_counts), 2) if paper_counts else 0,
                "advertisers_1_paper": sum(1 for n in paper_counts if n == 1),
                "advertisers_2_3_papers": sum(1 for n in paper_counts if 2 <= n <= 3),
                "advertisers_4_6_papers": sum(1 for n in paper_counts if 4 <= n <= 6),
                "advertisers_7_12_papers": sum(1 for n in paper_counts if 7 <= n <= 12),
                "advertisers_13_plus_papers": sum(1 for n in paper_counts if n >= 13),
            }
        )

    retention_matrix = []
    for row in advertiser_summary:
        fp = row["advertiser_fingerprint"]
        out = {
            "advertiser_fingerprint": fp,
            "advertiser_normalized": row["advertiser_normalized"],
            "active_months": row["active_months"],
            "retained_2024_to_2025": row["retained_2024_to_2025"],
            "max_papers_in_month": row["max_papers_in_month"],
        }
        active_months = {r["month"]: len({x["paper_slug"] for x in by_adv_month[(fp, r["month"])]}) for r in by_adv[fp]}
        for month in MONTHS:
            out[month] = active_months.get(month, 0)
        retention_matrix.append(out)

    cohort_rows = []
    first_seen_groups = defaultdict(list)
    for row in advertiser_summary:
        first_seen_groups[row["first_seen"]].append(row)
    for cohort_month in MONTHS:
        cohort = first_seen_groups.get(cohort_month, [])
        if not cohort:
            cohort_rows.append({"cohort_month": cohort_month, "new_advertisers": 0})
            continue
        out = {"cohort_month": cohort_month, "new_advertisers": len(cohort)}
        for offset in range(0, min(12, len(MONTHS) - MONTHS.index(cohort_month))):
            month = MONTHS[MONTHS.index(cohort_month) + offset]
            active_count = 0
            for adv in cohort:
                fp = adv["advertiser_fingerprint"]
                if (fp, month) in by_adv_month:
                    active_count += 1
            out[f"m+{offset}"] = active_count
            out[f"m+{offset}_rate"] = round(active_count / len(cohort), 3)
        cohort_rows.append(out)

    paper_month = []
    for coverage_row in coverage:
        key = (coverage_row["paper_slug"], coverage_row["month"])
        group = by_paper_month.get(key, [])
        paper_month.append(
            {
                **coverage_row,
                "paid_ad_rows": len(group),
                "unique_paid_advertisers": len({r["advertiser_fingerprint"] for r in group}),
                "full_page_or_back_rows": sum(1 for r in group if r["size_class"] in {"18FP", "BACK"}),
            }
        )

    review_rows = [r for r in rows if r["review_needed"] == "yes"]

    return {
        "placements": rows,
        "paid_rows": paid_rows,
        "advertiser_month": advertiser_month,
        "advertiser_summary": advertiser_summary,
        "monthly_summary": monthly_summary,
        "retention_matrix": retention_matrix,
        "cohorts": cohort_rows,
        "paper_month": paper_month,
        "coverage": coverage,
        "review_rows": review_rows,
    }


def write_outputs(data):
    with OUT_CSV.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(data["placements"][0].keys()))
        writer.writeheader()
        writer.writerows(data["placements"])

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    readme_rows = [
        ("Renna Media 2024-2025 Retention Analysis", ""),
        ("Scope", "All available 2024 and 2025 PDFs under inputs/editions."),
        ("Method", "Auto-detected ad-like boxes, phone fingerprints, and image-only full-page creative hashes."),
        ("Important Caveat", "This is not invoice data. Treat as a directional operating census until reconciled to customer exports."),
        ("Paid Rule", "Rollups include status=paid and status=paid_candidate, excluding Renna self, PSA, and editorial rows."),
        ("CSV", str(OUT_CSV.relative_to(ROOT))),
    ]
    for row in readme_rows:
        readme.append(row)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 110
    for cell in readme["A"]:
        cell.font = Font(bold=True)

    sheets = [
        ("Monthly Summary", data["monthly_summary"]),
        ("Advertiser Summary", data["advertiser_summary"]),
        ("Advertiser Month", data["advertiser_month"]),
        ("Retention Matrix", data["retention_matrix"]),
        ("Cohorts", data["cohorts"]),
        ("Paper Month", data["paper_month"]),
        ("Ad Placements", data["placements"]),
        ("Needs Review", data["review_rows"]),
    ]
    for name, rows in sheets:
        write_sheet(wb.create_sheet(name), rows)
    wb.save(OUT_XLSX)


def write_summary(data):
    monthly = data["monthly_summary"]
    adv = data["advertiser_summary"]
    paid_rows = data["paid_rows"]
    retained = [r for r in adv if r["retained_2024_to_2025"] == "yes"]
    active_12 = [r for r in adv if r["active_months"] >= 12]
    active_18 = [r for r in adv if r["active_months"] >= 18]
    top_reach = sorted(adv, key=lambda r: (-r["max_papers_in_month"], -r["active_months"]))[:15]
    avg_retention = [
        r["next_month_retention_rate"]
        for r in monthly
        if isinstance(r["next_month_retention_rate"], float)
    ]

    lines = [
        "# Advertiser Retention Analysis - 2024 and 2025",
        "",
        "## Outputs",
        "",
        "- `analysis/advertiser-retention-2024-2025.xlsx`",
        "- `analysis/advertiser-retention-2024-2025-placements.csv`",
        "- `analysis/build_2024_2025_retention_analysis.py`",
        "",
        "## Scope",
        "",
        f"- Issues scanned: {len(data['coverage'])}",
        f"- Paid/potential-paid placements: {len(paid_rows)}",
        f"- Paid/potential-paid advertiser fingerprints: {len(adv)}",
        f"- Rows requiring visual review: {len(data['review_rows'])}",
        "",
        "## Retention Signals",
        "",
        f"- Advertisers/fingerprints seen in both 2024 and 2025: {len(retained)}",
        f"- Advertisers/fingerprints active 12+ months: {len(active_12)}",
        f"- Advertisers/fingerprints active 18+ months: {len(active_18)}",
        f"- Average next-month retention across observed months: {round(sum(avg_retention) / len(avg_retention), 3) if avg_retention else 'n/a'}",
        "",
        "## Top Cross-Paper Buyers",
        "",
        "| Advertiser / Fingerprint | Active Months | Max Papers In Month | Avg Papers When Active | Primary Size |",
        "|---|---:|---:|---:|---|",
    ]
    for row in top_reach:
        lines.append(
            f"| {row['advertiser_normalized']} | {row['active_months']} | {row['max_papers_in_month']} | {row['avg_papers_when_active']} | {row['primary_size']} |"
        )
    lines.extend(
        [
            "",
            "## Caveats",
            "",
            "- This is a directional census built from PDFs, not a billing export.",
            "- Phone fingerprints are better for retention than image-only creative hashes. If a full-page advertiser changes creative, this may split one customer into multiple fingerprints.",
            "- Some PSAs/community notices can look like ads. They are filtered heuristically and retained in `Ad Placements` for audit.",
            "- Reconcile against Joe's customer/invoice export before using for valuation.",
        ]
    )
    SUMMARY_MD.write_text("\n".join(lines) + "\n")


def main():
    rows, coverage = scan()
    data = build_rollups(rows, coverage)
    write_outputs(data)
    write_summary(data)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"Issues scanned: {len(coverage)}")
    print(f"Inventory rows: {len(rows)}")
    print(f"Paid/potential paid rows: {len(data['paid_rows'])}")
    print(f"Advertiser fingerprints: {len(data['advertiser_summary'])}")
    print(f"Needs review rows: {len(data['review_rows'])}")


if __name__ == "__main__":
    main()
